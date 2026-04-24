"""
add_building_entry.py
---------------------
Automates the full DataTable configuration pipeline for a building facility or accessory entry.

New 4-step pipeline:
  1. DT_AchievementBuildMergeTag      — register achievement tag
  2. BuildingMergeAchievementList.xlsx — write row + precheckin (generates DT uasset)
  3. AQ DataAsset + DT patch           — create AQ asset, set AchievementQuery via update-datatable-row
  4. DT_BuildingBlockList              — add building block entry

Usage:
    python Tools/ConfigurationTools/ConfigFacilityOrAccessory/add_building_entry.py "Entity.FireFlyLamp" "萤火灯"
    python Tools/ConfigurationTools/ConfigFacilityOrAccessory/add_building_entry.py --batch batch_input.txt

Batch file format (one entry per line, comma-separated):
    Entity.FireFlyLamp,萤火灯
    Entity.Anvil,铁砧

Requirements:
    pip install requests openpyxl
Run from the project root (D:/Depot/Jun).
"""

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import textwrap
from pathlib import Path

try:
    import requests
except ImportError:
    print("ERROR: 'requests' library not found. Run: pip install requests")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ENTITY_TAG_MAPPING_XLSX = Path("Main/RawData/PLEntityTagMapping.xlsx")
ENTITY_TAG_MAPPING_CSV  = Path(
    "Main/AssetExports/DataTables/Game/003_DataTablePipeline/DT_PLEntityTagMapping.csv"
)
ACHIEVEMENT_LIST_XLSX   = Path("Main/RawData/AchievementList/BuildingMergeAchievementList.xlsx")
# Use the P4VUtils-installed copy (kept in sync with depot DevOps dist).
# Engine/Extras/... is the build output and may be stale / missing checks.
PRECHECKIN_EXE = Path(os.environ.get("LOCALAPPDATA", "")) / "Epic Games/P4VUtils/PreCheckin.exe"
INSTANCE_JSON           = Path("Main/.soft-ue-bridge/instance.json")

DT_ACHIEVEMENT_TAG       = "/Game/000_GlobalSettings/DataTables/DT_GameplayTag/DT_AchievementBuildMergeTag"
DT_ACHIEVEMENT_LIST      = "/Game/003_DataTablePipeline/AchievementList/DT_BuildingMergeAchievementList"
DT_BUILDING_BLOCK_LIST   = "/Game/000_GlobalSettings/DataTables/DT_BuildingBlockList"
AQ_ASSET_DIR             = "/Game/000_GlobalSettings/DataAssets/Achievement"

# Disk paths for p4 edit (relative to project root)
DT_ACHIEVEMENT_TAG_UASSET  = Path("Main/Content/000_GlobalSettings/DataTables/DT_GameplayTag/DT_AchievementBuildMergeTag.uasset")
DT_ACHIEVEMENT_LIST_UASSET = Path("Main/Content/003_DataTablePipeline/AchievementList/DT_BuildingMergeAchievementList.uasset")
DT_BUILDING_BLOCK_LIST_UASSET = Path("Main/Content/000_GlobalSettings/DataTables/DT_BuildingBlockList.uasset")

# EntityType.Building.Accessory tags that indicate SIZE (not type) — ignore these
ACCESSORY_SIZE_TAGS = {"Big", "Medium", "Small"}

# ---------------------------------------------------------------------------
# Bridge connection
# ---------------------------------------------------------------------------

def discover_bridge_url() -> str:
    if INSTANCE_JSON.exists():
        try:
            data = json.loads(INSTANCE_JSON.read_text(encoding="utf-8"))
            host = data.get("host", "127.0.0.1")
            port = data.get("port", 8080)
            return f"http://{host}:{port}/bridge"
        except Exception as e:
            print(f"  [WARN] Could not parse {INSTANCE_JSON}: {e}. Falling back to port 8080.")
    else:
        print(f"  [WARN] {INSTANCE_JSON} not found. Falling back to port 8080.")
    return "http://127.0.0.1:8080/bridge"


def bridge_call(bridge_url: str, tool_name: str, arguments: dict) -> dict:
    body = {
        "jsonrpc": "2.0",
        "id": "1",
        "method": "tools/call",
        "params": {"name": tool_name, "arguments": arguments},
    }
    resp = requests.post(bridge_url, json=body, timeout=60)
    resp.raise_for_status()
    result = resp.json()
    if "error" in result:
        raise RuntimeError(f"Bridge protocol error: {result['error']}")
    inner = result.get("result", {})
    if inner.get("isError"):
        text = inner.get("content", [{}])[0].get("text", "unknown error")
        raise RuntimeError(f"Tool error: {text}")
    return result


def bridge_call_add_row(bridge_url: str, tool_name: str, arguments: dict) -> str:
    """
    Like bridge_call but for add-datatable-row: returns 'added', 'skipped' (row exists), or raises.
    """
    try:
        bridge_call(bridge_url, tool_name, arguments)
        return "added"
    except RuntimeError as e:
        if "Row already exists" in str(e) or "already exists" in str(e).lower():
            return "skipped"
        raise

# ---------------------------------------------------------------------------
# Type resolution from DT_PLEntityTagMapping
# ---------------------------------------------------------------------------

def parse_gameplay_tag_names(raw: str) -> list[str]:
    """Extract all TagName values from a UE GameplayTagContainer CSV cell."""
    return re.findall(r'TagName="([^"]+)"', raw)


def _building_block_type_from_tags(tags: list[str]) -> str | None:
    """Derive BuildingBlock Type from a list of EntityType tag strings."""
    for tag in tags:
        if tag.startswith("EntityType.Building.Facility"):
            return "BuildingBlock.Facility"
    for tag in tags:
        if tag.startswith("EntityType.Building.Accessory."):
            sub = tag[len("EntityType.Building.Accessory."):]
            if sub not in ACCESSORY_SIZE_TAGS:
                return tag.replace("EntityType.Building", "BuildingBlock", 1)
    return None


def _resolve_from_xlsx(entity_tag: str) -> tuple[bool, str | None]:
    try:
        import openpyxl
    except ImportError:
        return False, None
    if not ENTITY_TAG_MAPPING_XLSX.exists():
        return False, None
    wb = openpyxl.load_workbook(ENTITY_TAG_MAPPING_XLSX, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        tag_col  = headers.index("EntityTag")
        type_col = headers.index("EntityTypes")
    except ValueError:
        wb.close()
        return False, None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[tag_col] == entity_tag:
            raw = row[type_col] or ""
            tags = [t.strip() for t in str(raw).split("|") if t.strip()]
            wb.close()
            return True, _building_block_type_from_tags(tags)
    wb.close()
    return False, None


def _resolve_from_csv(entity_tag: str) -> tuple[bool, str | None]:
    if not ENTITY_TAG_MAPPING_CSV.exists():
        return False, None
    with open(ENTITY_TAG_MAPPING_CSV, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("---", "").strip() != entity_tag:
                continue
            raw = row.get("EntityTypes", "")
            tags = parse_gameplay_tag_names(raw)
            return True, _building_block_type_from_tags(tags)
    return False, None


def resolve_building_block_type(entity_tag: str) -> str | None:
    found, result = _resolve_from_xlsx(entity_tag)
    if found:
        return result
    print("  [WARN] Not found in Excel. Trying CSV...")
    found, result = _resolve_from_csv(entity_tag)
    if found:
        return result
    raise ValueError(
        f"EntityTag '{entity_tag}' not found in Excel or CSV.\n"
        f"  Excel: {ENTITY_TAG_MAPPING_XLSX}\n"
        f"  CSV  : {ENTITY_TAG_MAPPING_CSV}"
    )

# ---------------------------------------------------------------------------
# P4 helpers
# ---------------------------------------------------------------------------

def p4_get_current_cl() -> int | None:
    """Return the most recent pending changelist for the current client, or None."""
    try:
        # Get current client name first
        info = subprocess.run(
            ["p4", "info"], capture_output=True, text=True, timeout=15
        )
        client_match = re.search(r"Client name:\s*(\S+)", info.stdout)
        client = client_match.group(1) if client_match else None

        cmd = ["p4", "changes", "-s", "pending", "-m", "1"]
        if client:
            cmd += ["-c", client]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
        m = re.search(r"Change\s+(\d+)", result.stdout)
        if m:
            return int(m.group(1))
    except Exception as e:
        print(f"  [WARN] p4 changes failed: {e}")
    return None


def p4_edit(path: Path, cl: int | None = None) -> bool:
    """Check out a file for edit, optionally into a specific CL."""
    cmd = ["p4", "edit"]
    if cl:
        cmd += ["-c", str(cl)]
    cmd.append(str(path))
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=15)
        output = result.stdout.strip() or result.stderr.strip()
        print(f"  p4 edit: {output}")
        return result.returncode == 0
    except Exception as e:
        print(f"  [WARN] p4 edit failed: {e}")
        return False

# ---------------------------------------------------------------------------
# Excel write — BuildingMergeAchievementList.xlsx
# ---------------------------------------------------------------------------

def write_excel_achievement_row(achievement_tag: str, chinese_name: str) -> str:
    """
    Append a new row to BuildingMergeAchievementList.xlsx.
    Returns 'added', 'skipped' (row with matching AchievementTag already exists), or raises.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed.")

    if not ACHIEVEMENT_LIST_XLSX.exists():
        raise RuntimeError(f"xlsx not found: {ACHIEVEMENT_LIST_XLSX}")

    dev_comment = f"建造：{chinese_name}"

    wb = openpyxl.load_workbook(ACHIEVEMENT_LIST_XLSX)
    ws = wb.active

    # Find last data row and max Id; also check for duplicate AchievementTag (col D)
    last_data_row = 3
    max_id = 0
    for row in ws.iter_rows(min_row=4, values_only=False):
        cell_id = row[0].value
        if cell_id is not None:
            last_data_row = row[0].row
            try:
                v = int(cell_id)
                if v > max_id:
                    max_id = v
            except (ValueError, TypeError):
                pass
        # col D (index 3) = AchievementTag
        if len(row) > 3 and row[3].value == achievement_tag:
            wb.close()
            return "skipped"

    new_id = max_id + 1
    insert_row = last_data_row + 1

    print(f"  Inserting at row {insert_row}, Id={new_id}")

    ws.cell(row=insert_row, column=1,  value=new_id)
    ws.cell(row=insert_row, column=2,  value=achievement_tag)
    ws.cell(row=insert_row, column=3,  value=dev_comment)
    ws.cell(row=insert_row, column=4,  value=achievement_tag)
    ws.cell(row=insert_row, column=5,  value=1)
    ws.cell(row=insert_row, column=6,  value=True)
    ws.cell(row=insert_row, column=7,  value="AssetTag.Achievement.Show.BuildMerge")
    ws.cell(row=insert_row, column=8,  value=1)
    ws.cell(row=insert_row, column=9,  value=0)
    ws.cell(row=insert_row, column=10, value=dev_comment)
    ws.cell(row=insert_row, column=11, value=dev_comment)
    ws.cell(row=insert_row, column=12, value=dev_comment)
    ws.cell(row=insert_row, column=13, value=None)
    ws.cell(row=insert_row, column=14, value=5)
    ws.cell(row=insert_row, column=15, value=1)
    ws.cell(row=insert_row, column=16, value=None)

    wb.save(ACHIEVEMENT_LIST_XLSX)
    wb.close()
    print(f"  [OK] Written row Id={new_id} to {ACHIEVEMENT_LIST_XLSX}")
    return "added"

# ---------------------------------------------------------------------------
# Precheckin
# ---------------------------------------------------------------------------

def run_precheckin(cl: int) -> bool:
    """Run the P4VUtils-installed PreCheckin.exe for the given changelist."""
    exe = PRECHECKIN_EXE
    if not exe.exists():
        print(f"  [ERROR] PreCheckin.exe not found: {exe}")
        return False

    depot_root = Path(".").resolve()
    print(f"  Running PreCheckin for CL {cl}...")
    try:
        result = subprocess.run(
            [str(exe), str(cl)],
            cwd=str(depot_root),
            timeout=300
        )
        if result.returncode != 0:
            print(f"  [WARN] PreCheckin.exe returned code {result.returncode}")
        return result.returncode == 0
    except subprocess.TimeoutExpired:
        print("  [ERROR] PreCheckin.exe timed out after 300s")
        return False
    except Exception as e:
        print(f"  [ERROR] PreCheckin.exe failed: {e}")
        return False

# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def process_entry(entity_tag: str, chinese_name: str, bridge_url: str) -> bool:
    """
    Run the full 4-step pipeline for one building entry.
    Returns True on success, False on failure.
    """
    print(f"\n{'='*60}")
    print(f"  EntityTag : {entity_tag}")
    print(f"  Name      : {chinese_name}")
    print(f"{'='*60}")

    if not entity_tag.startswith("Entity."):
        print(f"  [ERROR] EntityTag must start with 'Entity.' — got: {entity_tag}")
        return False

    suffix          = entity_tag[len("Entity."):]
    achievement_tag = f"Achievement.Show.BuildMerge.{suffix}"
    aq_name         = f"AQ_{suffix}"
    aq_game_path    = f"{AQ_ASSET_DIR}/{aq_name}"
    aq_full_ref     = f"{aq_game_path}.{aq_name}"

    print(f"  AchievementTag : {achievement_tag}")
    print(f"  AQ Asset       : {aq_game_path}")

    # -- Step 0: Resolve Type --
    print("\n[Step 0] Resolving BuildingBlock Type from DT_PLEntityTagMapping...")
    try:
        building_block_type = resolve_building_block_type(entity_tag)
    except Exception as e:
        print(f"  [ERROR] {e}")
        return False
    if building_block_type is None:
        print(f"  [WARN] No building type found. DT_BuildingBlockList Type will be empty.")
    else:
        print(f"  Type: {building_block_type}")

    # -- Step 1: DT_AchievementBuildMergeTag --
    print(f"\n[Step 1] Adding row to DT_AchievementBuildMergeTag...")
    cl = p4_get_current_cl()
    if cl:
        print(f"  Using CL: {cl}")
    else:
        print("  [WARN] No pending CL found.")

    p4_edit(DT_ACHIEVEMENT_TAG_UASSET, cl)
    try:
        status = bridge_call_add_row(bridge_url, "add-datatable-row", {
            "asset_path": DT_ACHIEVEMENT_TAG,
            "row_name": achievement_tag,
            "row_data": json.dumps({
                "Tag": achievement_tag,
                "DevComment": chinese_name,
            }),
        })
        print(f"  [{'OK' if status == 'added' else 'SKIP'}] {achievement_tag} ({status})")
    except Exception as e:
        print(f"  [ERROR] {e}")
        return False

    # Save to disk so precheckin's UE commandlet sees the new tag
    try:
        bridge_call(bridge_url, "save-asset", {"asset_path": DT_ACHIEVEMENT_TAG})
        print(f"  [OK] DT_AchievementBuildMergeTag saved")
    except Exception as e:
        print(f"  [ERROR] Save failed: {e}")
        return False

    # -- Step 2: Write xlsx + Precheckin (manages dt_metadata.json) --
    print(f"\n[Step 2] Writing BuildingMergeAchievementList.xlsx and running precheckin...")

    p4_edit(ACHIEVEMENT_LIST_XLSX, cl)

    try:
        xlsx_status = write_excel_achievement_row(achievement_tag, chinese_name)
    except Exception as e:
        print(f"  [ERROR] Failed to write xlsx row: {e}")
        return False

    if xlsx_status == "skipped":
        print(f"  [SKIP] Row already exists in xlsx — skipping export")
    elif cl:
        # Use Remote Control API to trigger excel_exporter INSIDE the running Editor.
        # This avoids file lock issues (no separate UE commandlet) and includes localization.
        print(f"  Triggering TriggerExcelPreCheckinExport via Remote Control (CL {cl})...")
        try:
            rc_resp = requests.put(
                "http://127.0.0.1:30010/remote/object/call",
                json={
                    "objectPath": "/Script/SoftUEBridgeEditor.Default__BridgeEditorFunctionLibrary",
                    "functionName": "TriggerExcelPreCheckinExport",
                    "parameters": {"Changelist": cl},
                },
                timeout=120,
            )
            rc_resp.raise_for_status()
            rc_data = rc_resp.json()
            ret_val = rc_data.get("ReturnValue", -1)
            if ret_val == 0:
                print(f"  [OK] TriggerExcelPreCheckinExport succeeded (ReturnValue=0)")
            else:
                print(f"  [WARN] TriggerExcelPreCheckinExport returned {ret_val} — check Editor log")
        except Exception as e:
            print(f"  [ERROR] Remote Control call failed: {e}")
            print(f"  [HINT] Is UE Editor running with Remote Control plugin enabled on port 30010?")
            return False
    else:
        print("  [WARN] Skipping export — no CL found.")

    # -- Step 3: Create AQ DataAsset + patch AchievementQuery --
    print(f"\n[Step 3] Creating AQ DataAsset and patching AchievementQuery...")

    py_script = textwrap.dedent(f"""\
        import unreal

        entity_tag_str  = "{entity_tag}"
        aq_name         = "AQ_{suffix}"
        aq_dir          = "{AQ_ASSET_DIR}"
        aq_game_path    = aq_dir + "/" + aq_name

        # --- Try to load an existing asset first (idempotent) ---
        aq = unreal.load_object(None, aq_game_path)

        if aq is None:
            f = unreal.DataAssetFactory()
            f.set_editor_property("data_asset_class", unreal.PLAchievementQuery)
            asset_tools = unreal.AssetToolsHelpers.get_asset_tools()
            aq = asset_tools.create_asset(aq_name, aq_dir, unreal.PLAchievementQuery, f)

        if aq is None:
            raise RuntimeError("Failed to create or load asset: " + aq_game_path)

        # --- Reset existing triggers (idempotent) ---
        aq.set_editor_property("counting_triggers", [])
        aq.set_editor_property("counting_additional_condition", None)

        # --- Build GameplayTagContainer with the EntityTag ---
        tag_obj = unreal.PLGameplayTagsFunctionLibrary.blueprint_request_gameplay_tag_from_name(entity_tag_str)
        tag_container = unreal.GameplayTagLibrary.make_gameplay_tag_container_from_tag(tag_obj)

        # --- Create PLAchievementTrigger_Think instanced sub-object ---
        trigger = unreal.new_object(unreal.PLAchievementTrigger_Think, aq)
        trigger.set_editor_property("entity_tags_to_check", tag_container)
        aq.set_editor_property("counting_triggers", [trigger])

        # --- Create PLAchievementCondition_BoolOr ---
        cond = unreal.new_object(unreal.PLAchievementCondition_BoolOr, aq)
        aq.set_editor_property("counting_additional_condition", cond)

        # --- Save ---
        unreal.EditorAssetLibrary.save_asset(aq_game_path, only_if_is_dirty=False)
        print("OK:" + aq_game_path)
    """)

    try:
        result = bridge_call(bridge_url, "run-python-script", {"script": py_script})
        output = str(result)
        if '"success": true' in output and "OK:" in output:
            print(f"  [OK] AQ asset: {aq_game_path}")
        else:
            err = re.search(r'"error":\s*"([^"]+)"', output)
            err_msg = err.group(1) if err else output[:300]
            print(f"  [ERROR] AQ script failed: {err_msg}")
            return False
    except Exception as e:
        print(f"  [ERROR] {e}")
        return False

    # Move AQ uasset from default CL into the working CL (Editor creates assets in default)
    if cl:
        aq_uasset = Path(f"Main/Content/000_GlobalSettings/DataAssets/Achievement/{aq_name}.uasset")
        try:
            r = subprocess.run(
                ["p4", "reopen", "-c", str(cl), str(aq_uasset)],
                capture_output=True, text=True, timeout=15,
            )
            print(f"  p4 reopen AQ: {r.stdout.strip() or r.stderr.strip()}")
        except Exception as e:
            print(f"  [WARN] p4 reopen AQ failed: {e}")

    # Patch AchievementQuery on the DT row created by precheckin
    print(f"  Patching AchievementQuery in DT_BuildingMergeAchievementList...")
    try:
        bridge_call(bridge_url, "update-datatable-row", {
            "asset_path": DT_ACHIEVEMENT_LIST,
            "row_name": achievement_tag,
            "row_data": json.dumps({
                "AchievementQuery": aq_full_ref,
            }),
        })
        print(f"  [OK] AchievementQuery set to {aq_full_ref}")
    except Exception as e:
        print(f"  [ERROR] Patching AchievementQuery failed: {e}")
        return False

    try:
        bridge_call(bridge_url, "save-asset", {"asset_path": DT_ACHIEVEMENT_LIST})
        print(f"  [OK] DT_BuildingMergeAchievementList saved")
    except Exception as e:
        print(f"  [ERROR] Save DT_BuildingMergeAchievementList failed: {e}")
        return False

    # -- Step 4: DT_BuildingBlockList --
    print(f"\n[Step 4] Adding row to DT_BuildingBlockList...")
    p4_edit(DT_BUILDING_BLOCK_LIST_UASSET, cl)
    row_data: dict = {
        "Desc": chinese_name,
        "BaseVariant": {"TagName": entity_tag},
    }
    if building_block_type:
        row_data["Type"] = {"TagName": building_block_type}
    try:
        status = bridge_call_add_row(bridge_url, "add-datatable-row", {
            "asset_path": DT_BUILDING_BLOCK_LIST,
            "row_name": entity_tag,
            "row_data": json.dumps(row_data),
        })
        print(f"  [{'OK' if status == 'added' else 'SKIP'}] {entity_tag} ({status})")
    except Exception as e:
        print(f"  [ERROR] {e}")
        return False

    try:
        bridge_call(bridge_url, "save-asset", {"asset_path": DT_BUILDING_BLOCK_LIST})
        print(f"  [OK] DT_BuildingBlockList saved")
    except Exception as e:
        print(f"  [ERROR] Save DT_BuildingBlockList failed: {e}")
        return False

    print(f"\n  All steps completed for '{entity_tag}' / '{chinese_name}'")
    return True

# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Add a building facility/accessory entry across DataTables via SoftUEBridge."
    )
    parser.add_argument("entity_tag", nargs="?", help="EntityTag, e.g. Entity.FireFlyLamp")
    parser.add_argument("chinese_name", nargs="?", help="Chinese display name, e.g. 萤火灯")
    parser.add_argument(
        "--batch",
        metavar="FILE",
        help="Batch mode: path to a CSV file with lines 'EntityTag,ChineseName'",
    )
    args = parser.parse_args()

    bridge_url = discover_bridge_url()
    print(f"SoftUEBridge URL: {bridge_url}")

    entries: list[tuple[str, str]] = []

    if args.batch:
        batch_file = Path(args.batch)
        if not batch_file.exists():
            print(f"ERROR: Batch file not found: {batch_file}")
            sys.exit(1)
        with open(batch_file, encoding="utf-8-sig", newline="") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                parts = line.split(",", 1)
                if len(parts) != 2:
                    print(f"[WARN] Skipping malformed line: {line!r}")
                    continue
                entries.append((parts[0].strip(), parts[1].strip()))
    elif args.entity_tag and args.chinese_name:
        entries.append((args.entity_tag, args.chinese_name))
    else:
        parser.print_help()
        sys.exit(1)

    failures = []
    for entity_tag, chinese_name in entries:
        ok = process_entry(entity_tag, chinese_name, bridge_url)
        if not ok:
            failures.append(entity_tag)

    print(f"\n{'='*60}")
    print(f"Done. {len(entries) - len(failures)}/{len(entries)} succeeded.")
    if failures:
        print(f"Failed entries: {', '.join(failures)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
